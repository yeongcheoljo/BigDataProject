from openpyxl import load_workbook
import pandas as pd
import glob

# Path to the directory with your files
file_path = "/content/*.xlsx"  # Update this path
all_files = sorted(glob.glob(file_path))  # Sort files alphabetically, ensuring Jan-Dec order

# List to collect data from each file
monthly_data_frames = []
for file in all_files:
    # Load workbook and select the active sheet
    wb = load_workbook(file)
    sheet = wb.active

    # Read data from each row
    data = []
    for row in sheet.iter_rows(values_only=True):
        # Check if any cell in the row is empty and replace it with the previous row's value
        if data and row[0] is None:
            # If current row's first cell is empty, use values from the last row in `data`
            last_row = data[-1]
            row = tuple(last_row[i] if cell is None else cell for i, cell in enumerate(row))
        data.append(row)

    # Convert data to a DataFrame and append to the list
    columns = data[0]  # First row as header
    df = pd.DataFrame(data[1:], columns=columns)
    monthly_data_frames.append(df)

# Concatenate all monthly DataFrames into one
combined_data = pd.concat(monthly_data_frames, ignore_index=True)

# Save the combined data to a new Excel file
combined_data.to_excel("노선별 이용량 23-combine.xlsx", index=False)
